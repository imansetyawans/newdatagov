import csv
import io
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Asset, Column, DQIssue


EMAIL_PATTERN = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
DATE_PATTERN = re.compile(r"^\d{4}-\d{2}-\d{2}$")
INTEGER_PATTERN = re.compile(r"^-?\d+$")
DECIMAL_PATTERN = re.compile(r"^-?\d+\.\d+$")
UUID_PATTERN = re.compile(r"^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}$")
URL_PATTERN = re.compile(r"^https?://\S+$", re.IGNORECASE)
MAX_UPLOAD_BYTES = 10 * 1024 * 1024


@dataclass(frozen=True)
class ParsedDataset:
    headers: list[str]
    rows: list[dict[str, Any]]


def normalize_identifier(value: str) -> str:
    normalized = re.sub(r"[^0-9a-zA-Z_]+", "_", value.strip())
    normalized = re.sub(r"_+", "_", normalized).strip("_")
    return normalized or "uploaded_dataset"


def parse_dataset_upload(filename: str, content: bytes) -> ParsedDataset:
    if len(content) > MAX_UPLOAD_BYTES:
        raise ValueError("File exceeds the 10 MB upload limit")

    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        return _parse_csv(content)
    if suffix == ".xlsx":
        return _parse_xlsx(content)
    raise ValueError("Unsupported file type. Upload a CSV or XLSX file")


def _parse_csv(content: bytes) -> ParsedDataset:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    headers = [header.strip() if header else "" for header in (reader.fieldnames or [])]
    _validate_headers(headers)
    rows = [
        {header: _clean_value(row.get(header)) for header in headers}
        for row in reader
    ]
    if not rows:
        raise ValueError("Uploaded dataset has no data rows")
    return ParsedDataset(headers=headers, rows=rows)


def _parse_xlsx(content: bytes) -> ParsedDataset:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise ValueError("XLSX upload support is not installed") from exc

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    row_iter = sheet.iter_rows(values_only=True)
    try:
        first_row = next(row_iter)
    except StopIteration as exc:
        raise ValueError("Uploaded workbook is empty") from exc

    headers = [str(value).strip() if value is not None else "" for value in first_row]
    _validate_headers(headers)
    rows = [
        {header: _clean_value(value) for header, value in zip(headers, row, strict=False)}
        for row in row_iter
        if any(value is not None and str(value).strip() for value in row)
    ]
    if not rows:
        raise ValueError("Uploaded dataset has no data rows")
    return ParsedDataset(headers=headers, rows=rows)


def _validate_headers(headers: list[str]) -> None:
    if not headers:
        raise ValueError("Uploaded dataset must include a header row")
    if any(not header for header in headers):
        raise ValueError("Column headers cannot be blank")
    normalized = [header.lower() for header in headers]
    if len(set(normalized)) != len(normalized):
        raise ValueError("Column headers must be unique")


def _clean_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def infer_data_type(values: list[Any]) -> str:
    usable = [str(value).strip() for value in values if value is not None and str(value).strip()]
    if not usable:
        return "TEXT"
    lowered = [value.lower() for value in usable]
    if all(value in {"true", "false", "0", "1", "yes", "no"} for value in lowered):
        return "BOOLEAN"
    if all(INTEGER_PATTERN.match(value) for value in usable):
        return "INTEGER"
    if all(INTEGER_PATTERN.match(value) or DECIMAL_PATTERN.match(value) for value in usable):
        return "DECIMAL"
    if all(DATE_PATTERN.match(value) for value in usable):
        return "DATE"
    return "TEXT"


def infer_standard_format(column_name: str, data_type: str, sample_values: list[Any]) -> str | None:
    values = [str(value).strip() for value in sample_values if value is not None and str(value).strip()]
    if not values:
        return None
    lowered_values = [value.lower() for value in values]
    lower_name = column_name.lower()
    lower_type = data_type.lower()

    if all(EMAIL_PATTERN.match(value) for value in values):
        return "valid email address"
    if all(DATE_PATTERN.match(value) for value in values):
        return "YYYY-MM-DD date"
    if all(UUID_PATTERN.match(value) for value in values):
        return "UUID"
    if all(URL_PATTERN.match(value) for value in values):
        return "URL"
    if all(value in {"true", "false", "0", "1", "yes", "no"} for value in lowered_values):
        return "boolean value"
    if all(INTEGER_PATTERN.match(value) for value in values):
        return "integer identifier" if lower_name == "id" or lower_name.endswith("_id") else "integer number"
    if all(DECIMAL_PATTERN.match(value) or INTEGER_PATTERN.match(value) for value in values):
        return "decimal number"
    if 1 < len(set(values)) <= 8 and all(len(value) <= 32 for value in values):
        return f"controlled vocabulary: {', '.join(values[:5])}"
    if all(value == value.lower() for value in values if any(character.isalpha() for character in value)):
        return "lowercase text"
    if "char" in lower_type or "text" in lower_type:
        return "free text"
    return None


def distinct_sample_values(values: list[Any], limit: int = 5) -> list[Any]:
    samples: list[Any] = []
    seen: set[str] = set()
    for value in values:
        if value is None or str(value).strip() == "":
            continue
        key = str(value)
        if key in seen:
            continue
        seen.add(key)
        samples.append(value)
        if len(samples) == limit:
            break
    return samples


def upsert_uploaded_asset(
    catalogue_db: Session,
    schema_name: str,
    table_name: str,
    table_description: str | None,
    parsed: ParsedDataset,
) -> Asset:
    source_path = f"upload.{schema_name}.{table_name}"
    scanned_at = datetime.utcnow()
    asset = catalogue_db.scalar(select(Asset).where(Asset.source_path == source_path))
    if asset is None:
        asset = Asset(
            connector_id=None,
            name=table_name,
            source_path=source_path,
            asset_type="table",
            schema_name=schema_name,
        )
        catalogue_db.add(asset)

    asset.description = table_description or asset.description
    asset.row_count = len(parsed.rows)
    asset.last_scanned_at = scanned_at
    asset.deleted_at = None

    existing_columns = {column.name: column for column in asset.columns}
    seen_columns: set[str] = set()
    for index, header in enumerate(parsed.headers):
        seen_columns.add(header)
        values = [row.get(header) for row in parsed.rows]
        samples = distinct_sample_values(values)
        data_type = infer_data_type(values)
        column = existing_columns.get(header)
        if column is None:
            column = Column(asset=asset, name=header, data_type=data_type)
            catalogue_db.add(column)
        column.data_type = data_type
        column.ordinal_position = index
        column.nullable = any(value is None for value in values)
        column.sample_values = samples
        column.standard_format = infer_standard_format(header, data_type, samples)

    for column_name, column in existing_columns.items():
        if column_name not in seen_columns:
            catalogue_db.delete(column)

    catalogue_db.flush()
    return asset


def calculate_uploaded_quality(
    catalogue_db: Session,
    quality_db: Session,
    asset: Asset,
    parsed: ParsedDataset,
) -> int:
    for issue in quality_db.scalars(select(DQIssue).where(DQIssue.asset_id == asset.id)).all():
        quality_db.delete(issue)

    column_by_name = {column.name: column for column in asset.columns}
    total = len(parsed.rows)
    issue_count = 0
    for header in parsed.headers:
        column = column_by_name[header]
        values = [row.get(header) for row in parsed.rows]
        non_empty = [value for value in values if value is not None and str(value).strip()]
        distinct = {str(value) for value in non_empty}

        column.completeness_score = _score(len(non_empty), total)
        column.uniqueness_score = _score(len(distinct), total)
        column.consistency_score = _consistency_score(column, values)
        column.accuracy_score = _accuracy_score(column, values)

        for metric_name, score in {
            "completeness": column.completeness_score,
            "uniqueness": column.uniqueness_score,
            "consistency": column.consistency_score,
            "accuracy": column.accuracy_score,
        }.items():
            if score is not None and score < 95:
                issue_count += 1
                quality_db.add(
                    DQIssue(
                        asset_id=asset.id,
                        column_id=column.id,
                        metric_name=metric_name,
                        severity="critical" if score < 70 else "warning",
                        status="open",
                        delta_value=None,
                        previous_score=None,
                        current_score=score,
                    )
                )

    asset.dq_score = _average(
        [
            score
            for column in asset.columns
            for score in [
                column.completeness_score,
                column.uniqueness_score,
                column.consistency_score,
                column.accuracy_score,
            ]
        ]
    )
    catalogue_db.flush()
    quality_db.flush()
    return issue_count


def _score(valid_count: int, total: int) -> float | None:
    if total == 0:
        return None
    return round((valid_count / total) * 100, 1)


def _average(values: list[float | None]) -> float | None:
    usable = [value for value in values if value is not None]
    if not usable:
        return None
    return round(sum(usable) / len(usable), 1)


def _consistency_score(column: Column, values: list[Any]) -> float | None:
    pattern = _pattern_for(column)
    if pattern is None:
        return 100.0
    usable = [value for value in values if value is not None and str(value).strip()]
    if not usable:
        return None
    return _score(sum(1 for value in usable if pattern.match(str(value).strip())), len(usable))


def _pattern_for(column: Column) -> re.Pattern | None:
    lower_name = column.name.lower()
    lower_type = column.data_type.lower()
    if "email" in lower_name:
        return EMAIL_PATTERN
    if "uuid" in lower_name:
        return UUID_PATTERN
    if "url" in lower_name or "uri" in lower_name:
        return URL_PATTERN
    if "date" in lower_name or "time" in lower_name or lower_type == "date":
        return DATE_PATTERN
    return None


def _accuracy_score(column: Column, values: list[Any]) -> float | None:
    usable = [value for value in values if value is not None and str(value).strip()]
    if not usable:
        return None
    return _score(sum(1 for value in usable if _accuracy_valid(column, value)), len(usable))


def _accuracy_valid(column: Column, value: Any) -> bool:
    lower_name = column.name.lower()
    if lower_name in {"age", "customer_age", "user_age"}:
        try:
            age = float(value)
        except (TypeError, ValueError):
            return False
        return 0 <= age <= 120
    if "income" in lower_name or "amount" in lower_name:
        try:
            return float(value) >= 0
        except (TypeError, ValueError):
            return False
    if "status" in lower_name:
        return str(value).lower() in {"active", "inactive", "pending", "closed", "open", "draft", "completed", "approved", "rejected"}
    return True
